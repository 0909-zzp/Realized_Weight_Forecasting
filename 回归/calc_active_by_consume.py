# -*- coding: utf-8 -*-
import pandas as pd
import openpyxl
from openpyxl.styles import Font

df = pd.read_excel(
    'D:/HuaweiMoveData/Users/27438/Desktop/笔记报表_补充后_monthly.xlsx',
    sheet_name=0
)

act_col = '去重后小红星站外活跃率（30日归因）'
consume_col = '消费'
read_col = '小红星任务期阅读UV'

for c in [consume_col, act_col, read_col]:
    df[c] = pd.to_numeric(df[c], errors='coerce')

corr_ac = df[[consume_col, act_col]].dropna().corr().iloc[0,1]
corr_ar = df[[read_col, act_col]].dropna().corr().iloc[0,1]
corr_cr = df[[consume_col, read_col]].dropna().corr().iloc[0,1]

print(f'消费 vs 活跃率 相关系数: {corr_ac:.4f}')
print(f'阅读UV vs 活跃率 相关系数: {corr_ar:.4f}')
print(f'消费 vs 阅读UV 相关系数: {corr_cr:.4f}')

print(f'\n消费非空: {df[consume_col].notna().sum()}/{len(df)}')
print(f'阅读UV非空: {df[read_col].notna().sum()}/{len(df)}')
print(f'活跃率非空: {df[act_col].notna().sum()}/{len(df)}')

df['年月'] = df['时间'].str[:7]

monthly_stats = []  # list of [年月, 笔记数, 总消费, 消费加权活跃率, 简单平均]
for ym, group in df.groupby('年月'):
    g = group[[consume_col, act_col]].dropna()
    if len(g) == 0 or g[consume_col].sum() == 0:
        continue
    total = g[consume_col].sum()
    weighted = (g[act_col] * g[consume_col]).sum() / total * 100
    simple = g[act_col].mean() * 100
    monthly_stats.append([str(ym), len(g), round(total, 2), round(weighted, 2), round(simple, 2)])

print('\n===== 各年月活跃率(消费加权) =====')
for row in monthly_stats:
    print(f"{row[0]} | {row[1]:>3}条 | 消费{row[2]:>10.0f}元 | "
          f"消费加权: {row[3]:>6.2f}% | 简单平均: {row[4]:>6.2f}%")

# 写入 Excel
wb = openpyxl.load_workbook('D:/HuaweiMoveData/Users/27438/Desktop/笔记报表_补充后_monthly.xlsx')
if 'Sheet4' in wb.sheetnames:
    del wb['Sheet4']
ws = wb.create_sheet('Sheet4')

headers = ['年月', '笔记数', '总消费(元)', '消费加权活跃率(%)', '简单平均活跃率(%)']
for ci, h in enumerate(headers, 1):
    ws.cell(row=1, column=ci, value=h).font = Font(bold=True)

for ri, row_data in enumerate(monthly_stats, 2):
    for ci, val in enumerate(row_data, 1):
        ws.cell(row=ri, column=ci, value=val)

ws.column_dimensions['A'].width = 14
ws.column_dimensions['B'].width = 10
ws.column_dimensions['C'].width = 14
ws.column_dimensions['D'].width = 20
ws.column_dimensions['E'].width = 20

wb.save('D:/HuaweiMoveData/Users/27438/Desktop/笔记报表_补充后_monthly.xlsx')
print('\n已保存到 Sheet4')
