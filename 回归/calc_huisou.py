import pandas as pd
import numpy as np
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, numbers
import os
import shutil

src = 'D:/HuaweiMoveData/Users/27438/Desktop/笔记报表_补充后_v3.xlsx'
dst = 'D:/HuaweiMoveData/Users/27438/Desktop/笔记报表_补充后_v4.xlsx'

df = pd.read_excel(src, sheet_name=0)
df_hr = df[df['回搜率'].notna()].copy()

groups = df_hr.groupby('时间')
results = []
for name, g in groups:
    n = len(g)
    imp_sum = g['展现量'].sum()
    spend_sum = g['消费'].sum()
    click_sum = g['点击量'].sum()
    huisou_num = g['回搜数量'].sum()
    rate_imp = np.average(g['回搜率'], weights=g['展现量'])
    rate_spend = np.average(g['回搜率'], weights=g['消费'].clip(lower=0.01))
    rate_simple = g['回搜率'].mean()
    results.append({
        '时间段': name,
        '笔记数': int(n),
        '总展现量': int(imp_sum),
        '总点击量': int(click_sum),
        '总消费': round(spend_sum, 2),
        '总回搜数量': int(huisou_num),
        '综合回搜率(展现量加权)': round(rate_imp, 6),
        '综合回搜率(消费加权)': round(rate_spend, 6),
        '简单平均回搜率(不推荐)': round(rate_simple, 6)
    })

result_df = pd.DataFrame(results).sort_values('时间段')

# 先复制原文件
shutil.copy(src, dst)

wb = load_workbook(dst)
# 如果已存在同名sheet，删除重建
if '各时段综合回搜率' in wb.sheetnames:
    del wb['各时段综合回搜率']

ws = wb.create_sheet('各时段综合回搜率', 1)

# 表头样式
header_fill = PatternFill('solid', fgColor='4472C4')
header_font = Font(bold=True, color='FFFFFF', size=11)
pct_fmt = '0.00%'
num_fmt = '#,##0'

headers = list(result_df.columns)
for col_idx, h in enumerate(headers, 1):
    cell = ws.cell(row=1, column=col_idx, value=h)
    cell.fill = header_fill
    cell.font = header_font
    cell.alignment = Alignment(horizontal='center')

# 写入数据
for row_idx, row in result_df.iterrows():
    for col_idx, col_name in enumerate(headers):
        val = row[col_name]
        cell = ws.cell(row=row_idx + 2, column=col_idx + 1, value=val)
        if '回搜率' in col_name:
            cell.number_format = pct_fmt
        elif col_name in ['总展现量', '总点击量', '总回搜数量']:
            cell.number_format = num_fmt
        elif col_name == '总消费':
            cell.number_format = '#,##0.00'

# 列宽
col_widths = {'时间段': 22, '笔记数': 10, '总展现量': 14, '总点击量': 12,
              '总消费': 14, '总回搜数量': 14, '综合回搜率(展现量加权)': 24,
              '综合回搜率(消费加权)': 24, '简单平均回搜率(不推荐)': 24}
for col_idx, h in enumerate(headers, 1):
    ws.column_dimensions[ws.cell(row=1, column=col_idx).column_letter].width = col_widths.get(h, 12)

wb.save(dst)
print(f'已保存至: {dst}')

pd.set_option('display.float_format', lambda x: f'{x:.4f}' if abs(x) < 100 and abs(x) > 0.01 else f'{x:.6f}')
pd.set_option('display.max_columns', 10)
pd.set_option('display.width', 300)
print(result_df.to_string(index=False))
print(f'\n共 {len(result_df)} 个时间段')
