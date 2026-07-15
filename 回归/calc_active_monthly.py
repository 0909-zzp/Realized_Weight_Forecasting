import pandas as pd
import numpy as np
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment

desktop = 'D:/HuaweiMoveData/Users/27438/Desktop'
src = f'{desktop}/笔记报表_补充后_monthly.xlsx'

df = pd.read_excel(src, sheet_name=0)
df['年月'] = df['时间'].str[:7]

# 使用笔记报表版（数据更全，356/388）
rate_col = '小红星站外活跃率(去重后30日归因)'
uv_col = '小红星站外活跃行为UV(去重后30日归因)'
cost_col = '小红星站外活跃成本(去重后30日归因)'
readuv_col = '小红星任务期阅读UV'

df_sub = df[[rate_col, uv_col, readuv_col, cost_col, '年月', '展现量', '消费']].copy()
for c in [rate_col, uv_col, readuv_col, cost_col]:
    df_sub[c] = pd.to_numeric(df_sub[c], errors='coerce')
df_act = df_sub[df_sub[rate_col].notna()].copy()

groups = df_act.groupby('年月')
results = []
for name, g in groups:
    n = len(g)
    readuv_sum = g[readuv_col].sum()
    uv_sum = g[uv_col].sum()
    spend_sum = g['消费'].sum()
    # 阅读UV加权
    rate_readuv = np.average(g[rate_col], weights=g[readuv_col].clip(lower=1))
    # 展现量加权（备选）
    rate_imp = np.average(g[rate_col], weights=g['展现量'].clip(lower=1))
    # 简单平均
    rate_simple = g[rate_col].mean()
    results.append({
        '年月': name,
        '笔记数': int(n),
        '总阅读UV': int(readuv_sum),
        '总活跃UV': int(uv_sum),
        '总消费': round(spend_sum, 2),
        '综合活跃率(阅读UV加权)': round(rate_readuv, 6),
        '综合活跃率(展现量加权备选)': round(rate_imp, 6),
        '简单平均活跃率(不推荐)': round(rate_simple, 6)
    })

result_df = pd.DataFrame(results).sort_values('年月')

wb = load_workbook(src)
sheet_name = '各年月活跃率'
if sheet_name in wb.sheetnames:
    del wb[sheet_name]

ws = wb.create_sheet(sheet_name, 2)

header_fill = PatternFill('solid', fgColor='ED7D31')
header_font = Font(bold=True, color='FFFFFF', size=11)

headers = list(result_df.columns)
for col_idx, h in enumerate(headers, 1):
    cell = ws.cell(row=1, column=col_idx, value=h)
    cell.fill = header_fill
    cell.font = header_font
    cell.alignment = Alignment(horizontal='center')

for row_idx, row in result_df.iterrows():
    for col_idx, col_name in enumerate(headers):
        val = row[col_name]
        cell = ws.cell(row=row_idx + 2, column=col_idx + 1, value=val)
        if '活跃率' in col_name:
            cell.number_format = '0.00%'
        elif col_name in ['总阅读UV', '总活跃UV']:
            cell.number_format = '#,##0'
        elif col_name == '总消费':
            cell.number_format = '#,##0.00'

col_widths = {'年月': 12, '笔记数': 10, '总阅读UV': 14, '总活跃UV': 14,
              '总消费': 14, '综合活跃率(阅读UV加权)': 24,
              '综合活跃率(展现量加权备选)': 24, '简单平均活跃率(不推荐)': 24}
for col_idx, h in enumerate(headers, 1):
    ws.column_dimensions[ws.cell(row=1, column=col_idx).column_letter].width = col_widths.get(h, 12)

wb.save(src)
print(f'已保存至: {src}')

pd.set_option('display.float_format', lambda x: f'{x:.4f}' if abs(x) < 100 and abs(x) > 0.001 else f'{x:.6f}')
pd.set_option('display.max_columns', 10)
pd.set_option('display.width', 300)
print(result_df.to_string(index=False))
print(f'\n共 {len(result_df)} 个月份，已写入工作表"{sheet_name}"')
