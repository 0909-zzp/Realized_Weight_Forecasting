# -*- coding: utf-8 -*-
import pandas as pd
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

df = pd.read_excel('C:/Users/27438/Downloads/全部数据 (2).xlsx', sheet_name=0)

# 指标列
spend_col = '投放花费'
cpm_col = 'CPM'
search_count_col = '看后搜次数'
search_rate_col = '看后搜率'

# 转数值
for c in [spend_col, search_count_col, search_rate_col]:
    df[c] = pd.to_numeric(df[c], errors='coerce')

# CPM列可能有"元"后缀，清理
df[cpm_col] = df[cpm_col].astype(str).str.replace('元', '').str.strip()
df[cpm_col] = pd.to_numeric(df[cpm_col], errors='coerce')

print(f'总行数: {len(df)}')
for c in [spend_col, cpm_col, search_count_col, search_rate_col]:
    valid = df[c].notna().sum()
    print(f'{c}: {valid}/{len(df)} 非空')

# 提取年月 - 使用视频实际发布时间
df['年月'] = pd.to_datetime(df['视频实际发布时间']).dt.strftime('%Y-%m')
print(f'\n年月分布:')
print(df['年月'].value_counts().sort_index())

# 按月汇总
results = []
for ym, group in df.groupby('年月'):
    g = group[[spend_col, cpm_col, search_count_col, search_rate_col]].dropna(
        subset=[spend_col], how='any')
    
    total_spend = g[spend_col].sum()
    if total_spend == 0:
        continue
    
    n = len(g)
    
    # CPM - 投放花费加权 (等同于 total_spend / total_impressions * 1000)
    cpm_valid = g[[cpm_col, spend_col]].dropna()
    if len(cpm_valid) > 0 and cpm_valid[spend_col].sum() > 0:
        cpm_weighted = (cpm_valid[cpm_col] * cpm_valid[spend_col]).sum() / cpm_valid[spend_col].sum()
        cpm_simple = cpm_valid[cpm_col].mean()
    else:
        cpm_weighted = None
        cpm_simple = None
    
    # 看后搜次数 - 投放花费加权
    sc_valid = g[[search_count_col, spend_col]].dropna()
    if len(sc_valid) > 0 and sc_valid[spend_col].sum() > 0:
        sc_weighted = (sc_valid[search_count_col] * sc_valid[spend_col]).sum() / sc_valid[spend_col].sum()
        sc_simple = sc_valid[search_count_col].mean()
    else:
        sc_weighted = None
        sc_simple = None
    
    # 看后搜率 - 投放花费加权
    sr_valid = g[[search_rate_col, spend_col]].dropna()
    if len(sr_valid) > 0 and sr_valid[spend_col].sum() > 0:
        sr_weighted = (sr_valid[search_rate_col] * sr_valid[spend_col]).sum() / sr_valid[spend_col].sum()
        sr_simple = sr_valid[search_rate_col].mean()
    else:
        sr_weighted = None
        sr_simple = None
    
    results.append({
        '年月': ym,
        '笔记数': n,
        '总投放花费(元)': round(total_spend, 2),
        'CPM(消费加权)': round(cpm_weighted, 2) if cpm_weighted is not None else '-',
        'CPM(简单平均)': round(cpm_simple, 2) if cpm_simple is not None else '-',
        '看后搜次数(消费加权)': round(sc_weighted, 2) if sc_weighted is not None else '-',
        '看后搜次数(简单平均)': round(sc_simple, 2) if sc_simple is not None else '-',
        '看后搜率(消费加权%)': round(sr_weighted * 100, 4) if sr_weighted is not None else '-',
        '看后搜率(简单平均%)': round(sr_simple * 100, 4) if sr_simple is not None else '-',
    })

result_df = pd.DataFrame(results)

print('\n===== 各年月投放花费加权综合指标 =====')
for _, row in result_df.iterrows():
    print(f"\n{row['年月']} | {row['笔记数']}条 | 总花费{row['总投放花费(元)']:.0f}元")
    print(f"  CPM:          消费加权={row['CPM(消费加权)']}元 | 简单平均={row['CPM(简单平均)']}元")
    print(f"  看后搜次数:    消费加权={row['看后搜次数(消费加权)']} | 简单平均={row['看后搜次数(简单平均)']}")
    print(f"  看后搜率:      消费加权={row['看后搜率(消费加权%)']}% | 简单平均={row['看后搜率(简单平均%)']}%")

# 写入新的Excel文件
output_path = 'D:/HuaweiMoveData/Users/27438/Desktop/指标汇总_投放花费加权.xlsx'

with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
    # Sheet1: 原始数据
    df.drop(columns=['年月'], inplace=False).to_excel(writer, sheet_name='原始数据', index=False)
    # Sheet2: 汇总
    result_df.to_excel(writer, sheet_name='汇总_投放花费加权', index=False)

# 格式美化
from openpyxl.utils import get_column_letter
wb = openpyxl.load_workbook(output_path)
ws = wb['汇总_投放花费加权']

header_fill = PatternFill(start_color='E8833A', end_color='E8833A', fill_type='solid')
header_font = Font(bold=True, color='FFFFFF', size=11)

for ci in range(1, ws.max_column + 1):
    cell = ws.cell(row=1, column=ci)
    cell.fill = header_fill
    cell.font = header_font
    cell.alignment = Alignment(horizontal='center')
    ws.column_dimensions[get_column_letter(ci)].width = 20

wb.save(output_path)
print(f'\n已保存: {output_path}')
